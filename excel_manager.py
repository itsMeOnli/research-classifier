import openpyxl

class ExcelManager:
    def fill_row(self, ws, row, data, analysis):
        ws[f"A{row}"] = data["title"]
        ws[f"B{row}"] = data["year"]
        ws[f"E{row}"] = data["authors"]
        ws[f"F{row}"] = data["url"]

        if analysis:
            ws[f"H{row}"] = analysis["objectives"]
            ws[f"I{row}"] = analysis["findings"]

            col = 10
            for section in ["alignment", "national_outcomes"]:
                for value in analysis[section].values():
                    ws.cell(row=row, column=col).value = value
                    col += 1
